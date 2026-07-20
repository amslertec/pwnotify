"""Security A1 (CWE-1236): CSV/XLSX formula injection in the user export.

`display_name`, `upn`, `mail`, `other_mails`, `department`, `job_title` all originate from
the Entra sync, i.e. from attributes a helpdesk agent, a self-service user or a compromised
account can write. A value with a leading formula trigger (`=`, `+`, `-`, `@`, tab, CR) is
persisted by openpyxl as `data_type='f'` (a REAL formula that Excel evaluates on open) and is
merely quoted -- not neutralised -- by `csv.writer`, so Excel still evaluates it. The export
must neutralise every string cell so spreadsheet apps treat it as text.

These tests exercise the pure builder functions directly (no DB): RED against the old
builders (`data_type == 'f'` / an evaluable CSV formula), GREEN once each string cell that
starts with a trigger is prefixed with an apostrophe.
"""

from __future__ import annotations

import csv
import io

from app.api.routes.users import _build_csv, _build_xlsx
from openpyxl import load_workbook

_HEADERS = ["displayName", "daysLeft", "expiryDate"]


def test_xlsx_leading_equals_is_stored_as_text_not_formula() -> None:
    # "=1+1" with the old builder lands as data_type='f' (evaluated by Excel).
    buf = _build_xlsx(_HEADERS, [["=1+1", 7, "2026-07-20"]])
    wb = load_workbook(io.BytesIO(buf.getvalue()))
    cell = wb["Users"].cell(row=2, column=1)

    assert cell.data_type == "s", "leading '=' must be stored as a string, not a formula"
    # openpyxl reads the raw stored text, which keeps the neutralising apostrophe; Excel
    # hides that apostrophe on display and shows the literal "=1+1" without evaluating it.
    assert cell.value == "'=1+1"


def test_xlsx_all_triggers_are_neutralised() -> None:
    for trigger in ("=", "+", "-", "@", "\t", "\r"):
        payload = f"{trigger}cmd|'/c calc'!A0"
        buf = _build_xlsx(_HEADERS, [[payload, 1, "2026-07-20"]])
        wb = load_workbook(io.BytesIO(buf.getvalue()))
        cell = wb["Users"].cell(row=2, column=1)
        assert cell.data_type == "s", f"trigger {trigger!r} must stay text, not formula"


def test_csv_hyperlink_is_neutralised() -> None:
    payload = '=HYPERLINK("http://evil","x")'
    out = _build_csv(_HEADERS, [[payload, 7, "2026-07-20"]])
    rows = list(csv.reader(io.StringIO(out)))
    data = rows[1]

    # Old builder writes the raw formula (Excel evaluates it); the fix prepends an apostrophe.
    assert data[0].startswith("'="), "CSV formula cell must be prefixed with an apostrophe"
    assert data[0] == "'" + payload


def test_legit_string_and_typed_cells_pass_through() -> None:
    # A benign name and typed numeric/date cells must be untouched (no accidental quoting).
    buf = _build_xlsx(_HEADERS, [["Erika Mustermann", 7, "2026-07-20"]])
    wb = load_workbook(io.BytesIO(buf.getvalue()))
    ws = wb["Users"]
    assert ws.cell(row=2, column=1).value == "Erika Mustermann"
    assert ws.cell(row=2, column=2).value == 7  # int stays typed
    assert ws.cell(row=2, column=3).value == "2026-07-20"  # ISO date starts with a digit

    out = _build_csv(_HEADERS, [["Erika Mustermann", 7, "2026-07-20"]])
    data = list(csv.reader(io.StringIO(out)))[1]
    assert data == ["Erika Mustermann", "7", "2026-07-20"]
